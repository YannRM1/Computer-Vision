"""
PROGRAMME 2 : Lecture automatique des formulaires PDF

Fonctions principales :
  autoReadForm(pdf_dir, signatures_dir, results_dir)
  autoReadFormID(pdf_path, desc_db, results_dir)

Produit : un fichier XLSX par PDF, avec deux onglets :
  - PAGE-01 : infos de la page 1 (identité, conditions, notes)
  - EXAM    : réponses aux questions (choix, mantisse, exposant, unité)
"""

import os
import cv2
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

from utils.pdf_utils      import pdf_to_images
from utils.grid_decoder   import normalize_page, extract_cryptogram
from utils.page1_parser   import parse_page1
from utils.exam_parser    import parse_exam_pages, questions_to_exam_rows, CHOICE_COLS
from utils.signature_utils import load_signatures, build_descriptor_db

# Index de la première page d'examen (0-indexé) = page 5 dans les PDFs
EXAM_START_PAGE = 4
# DPI pour la conversion PDF -> image
PDF_DPI = 150


# ---------------------------------------------------------------------------
# Écriture du fichier xlsx résultat
# ---------------------------------------------------------------------------

def _write_xlsx(page1_data: dict,
                exam_rows: list[dict],
                xlsx_path: str) -> None:
    """
    Crée et enregistre le fichier xlsx avec deux onglets.
    Structure PAGE-01 :
      Colonne A = nom du champ, Colonne B = valeur
    """
    wb = Workbook()

    # ---- Onglet PAGE-01 --------------------------------------------------
    ws1 = wb.active
    ws1.title = "PAGE-01"

    page1_fields = [
        ("Module",                  page1_data.get("module")),
        ("Professor",               page1_data.get("professor")),
        ("Date",                    page1_data.get("date")),
        ("Code",                    page1_data.get("code")),
        ("Notes de cours",          page1_data.get("notes_cours")),
        ("Notes manuscrites",       page1_data.get("notes_manuscrites")),
        ("Ordinateur portable",     page1_data.get("ordinateur")),
        ("Calculatrice ",           page1_data.get("calculatrice")),
        ("Feuilles brouillon",      page1_data.get("brouillon")),
        ("Note maximale",           page1_data.get("note_max")),
        ("Note pour valider",       page1_data.get("note_valid")),
        (None,                      None),
        ("Prénom",                  page1_data.get("prenom")),
        ("Nom",                     page1_data.get("nom")),
        ("Validation signature",    page1_data.get("validation_signature")),
        ("Group",                   page1_data.get("group")),
        ("STUDENT ID",              page1_data.get("student_id")),
        ("Validation cryptogramme", page1_data.get("validation_cryptogramme")),
    ]
    for row_idx, (label, value) in enumerate(page1_fields, start=1):
        ws1.cell(row=row_idx, column=1, value=label)
        ws1.cell(row=row_idx, column=2, value=value)

    # ---- Onglet EXAM -----------------------------------------------------
    ws2 = wb.create_sheet(title="EXAM")

    headers = ["QUESTION"] + CHOICE_COLS + ["MANTISSE", "EXPOSANT", "UNITE"]
    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for row_idx, q_row in enumerate(exam_rows, start=2):
        ws2.cell(row=row_idx, column=1, value=q_row["QUESTION"])
        for c_idx, col in enumerate(CHOICE_COLS, start=2):
            ws2.cell(row=row_idx, column=c_idx, value=q_row.get(col))
        ws2.cell(row=row_idx, column=10, value=q_row.get("MANTISSE"))
        ws2.cell(row=row_idx, column=11, value=q_row.get("EXPOSANT"))
        ws2.cell(row=row_idx, column=12, value=q_row.get("UNITE"))

    wb.save(xlsx_path)


# ---------------------------------------------------------------------------
# Traitement d'un PDF (sous-fonction)
# ---------------------------------------------------------------------------

def autoReadFormID(pdf_path: str,
                   desc_db: dict,
                   results_dir: str) -> str:
    """
    Lit un formulaire PDF et génère le xlsx correspondant.

    Args:
        pdf_path    : chemin vers le PDF
        desc_db     : base de descripteurs de signatures
        results_dir : répertoire de sortie

    Returns:
        Chemin vers le xlsx généré.
    """
    pdf_name     = os.path.splitext(os.path.basename(pdf_path))[0]
    xlsx_path    = os.path.join(results_dir, pdf_name + ".xlsx")

    print(f"  [PDF] {os.path.basename(pdf_path)}", end=" ", flush=True)

    # 1. Convertir toutes les pages en images
    images = pdf_to_images(pdf_path, dpi=PDF_DPI)
    if not images:
        print("-> 0 pages !")
        return xlsx_path

    # 2. Extraire les cryptogrammes de toutes les pages (sauf page 1)
    crypto_pages = []
    for pg_img in images[1:]:
        norm = normalize_page(pg_img)
        c = extract_cryptogram(norm)
        if c is not None and c.size > 0:
            crypto_pages.append(c)

    # 3. Parser la page 1
    page1_data = parse_page1(
        images[0],
        desc_db=desc_db,
        crypto_pages=crypto_pages,
    )

    # 4. Parser les pages d'examen
    questions     = parse_exam_pages(images, exam_start_page=EXAM_START_PAGE)
    exam_rows     = questions_to_exam_rows(questions)

    # 5. Écrire le xlsx
    _write_xlsx(page1_data, exam_rows, xlsx_path)

    print(f"-> {len(questions)} questions, student_id={page1_data.get('student_id')}")
    return xlsx_path


# ---------------------------------------------------------------------------
# Fonction principale Programme 2
# ---------------------------------------------------------------------------

def autoReadForm(pdf_dir: str,
                 signatures_dir: str,
                 results_dir: str) -> list[str]:
    """
    Lit tous les formulaires PDF d'un répertoire.

    Args:
        pdf_dir        : répertoire contenant les fichiers PDF
        signatures_dir : répertoire / ZIP des signatures
        results_dir    : répertoire de sortie

    Returns:
        Liste des chemins xlsx générés.
    """
    os.makedirs(results_dir, exist_ok=True)

    # Charger la base de signatures
    print(f"[P2] Chargement des signatures depuis : {signatures_dir}")
    raw_db  = load_signatures(signatures_dir)
    print(f"  -> {len(raw_db)} eleves, {sum(len(v) for v in raw_db.values())} signatures")
    desc_db = build_descriptor_db(raw_db)

    # Lister les PDFs
    pdfs = sorted([
        f for f in os.listdir(pdf_dir)
        if f.lower().endswith(".pdf")
    ])
    print(f"[P2] {len(pdfs)} formulaires PDF à traiter dans : {pdf_dir}")

    generated = []
    for pdf_name in pdfs:
        pdf_path = os.path.join(pdf_dir, pdf_name)
        xlsx_path = autoReadFormID(pdf_path, desc_db, results_dir)
        generated.append(xlsx_path)

    print(f"[P2] Terminé. {len(generated)} fichiers générés dans : {results_dir}")
    return generated


# ---------------------------------------------------------------------------
# Point d'entrée direct
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    if len(sys.argv) == 4:
        autoReadForm(sys.argv[1], sys.argv[2], sys.argv[3])
    else:
        print("Usage: python autoReadForm.py <pdf_dir> <sig_dir> <results_dir>")
