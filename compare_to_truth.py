"""
Compare les xlsx produits par ton pipeline aux xlsx vérités terrain.

Vérité terrain  : PROJECT 2026 -DATABASE-20260518/FORMX/EXAM_FORMX_NNNNN.xlsx
                  (fourni par les profs, à côté de chaque pdf/photo)
Production      : EXAM_FORMX_RESULTS/EXAM_FORMX_NNNNN.xlsx
                  (généré par autoReadForm)

Pour chaque xlsx :
  - Onglet PAGE-01 : 18 lignes (Module, Professor, Date, ..., STUDENT ID)
  - Onglet EXAM    : grille des réponses (CHOIX A-H, MANTISSE, EXPOSANT, UNITE)

Sortie :
  - Récap par axe (imprimé / manuscrit / graphique / signature)
  - CSV détaillé : compare_results.csv  (champ par champ, OK/KO/MISSING)

Usage :
  python compare_to_truth.py [chemin_DATABASE] [chemin_RESULTS]
"""

import os
import sys
import csv
import re
from datetime import datetime

import openpyxl


# Mapping ligne PAGE-01 -> axe d'évaluation (cf. §6 du sujet)
PAGE01_AXES = {
    1:  ("Module",                  "imprime"),
    2:  ("Professor",               "imprime"),
    3:  ("Date",                    "imprime"),
    4:  ("Code",                    "imprime"),
    5:  ("Notes de cours",          "graphique"),
    6:  ("Notes manuscrites",       "graphique"),
    7:  ("Ordinateur portable",     "graphique"),
    8:  ("Calculatrice",            "graphique"),
    9:  ("Feuilles brouillon",      "graphique"),
    10: ("Note maximale",           "imprime"),
    11: ("Note pour valider",       "imprime"),
    13: ("Prenom",                  "manuscrit"),
    14: ("Nom",                     "manuscrit"),
    15: ("Validation signature",    "signature"),
    16: ("Group",                   "graphique"),
    17: ("STUDENT ID",              "graphique"),
    18: ("Validation cryptogramme", "graphique"),
}


def normalize_value(v):
    """Normalise pour comparaison (date->str, str->lower trimmed, etc.)."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        return v.strip().lower()
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def cells_equal(a, b):
    na = normalize_value(a)
    nb = normalize_value(b)
    if na is None and nb is None:
        return True
    if na is None or nb is None:
        return False
    if isinstance(na, (int, float)) and isinstance(nb, (int, float)):
        return abs(na - nb) < 1e-3
    return na == nb


def compare_page1(prod_ws, truth_ws):
    """Retourne liste de tuples (ligne, libellé, axe, truth, prod, ok)."""
    rows = []
    for r, (label, axis) in PAGE01_AXES.items():
        t = truth_ws.cell(r, 2).value
        p = prod_ws.cell(r, 2).value
        ok = cells_equal(t, p)
        rows.append((r, label, axis, t, p, ok))
    return rows


def compare_exam(prod_ws, truth_ws):
    """Compare l'onglet EXAM par cellule. Retourne (cellule, axe, truth, prod, ok)."""
    rows = []
    # Headers: col 1=QUESTION, col 2-9=CHOIX A-H, col 10=MANTISSE, col 11=EXPOSANT, col 12=UNITE
    axis_map = {1: "imprime", 10: "manuscrit", 11: "manuscrit", 12: "manuscrit"}
    for c in range(2, 10): axis_map[c] = "graphique"  # CHOIX A-H

    max_r = max(truth_ws.max_row, prod_ws.max_row)
    for r in range(2, max_r + 1):
        # Skip empty truth rows
        if truth_ws.cell(r, 1).value is None:
            continue
        for c in range(1, 13):
            t = truth_ws.cell(r, c).value
            p = prod_ws.cell(r, c).value if r <= prod_ws.max_row else None
            if t is None and p is None:
                continue
            axis = axis_map.get(c, "autre")
            ok = cells_equal(t, p)
            cell = f"R{r}C{c}"
            rows.append((cell, axis, t, p, ok))
    return rows


def find_pairs(data_root, results_root):
    """Trouve les paires (vérité, production) à comparer."""
    pairs = []
    for form in ("FORM1", "FORM2", "FORM3"):
        truth_dir   = os.path.join(data_root, form)
        results_dir = os.path.join(results_root, f"EXAM_{form}_RESULTS")
        if not os.path.isdir(truth_dir):
            continue
        for f in sorted(os.listdir(truth_dir)):
            m = re.match(rf"EXAM_{form}_(\d+)\.xlsx", f)
            if not m: continue
            truth_path = os.path.join(truth_dir, f)
            prod_path  = os.path.join(results_dir, f)
            if os.path.isfile(prod_path):
                pairs.append((form, m.group(1), truth_path, prod_path))
    return pairs


def main():
    data_root    = sys.argv[1] if len(sys.argv) > 1 else "PROJECT 2026 -DATABASE-20260518"
    results_root = sys.argv[2] if len(sys.argv) > 2 else "."

    pairs = find_pairs(data_root, results_root)
    if not pairs:
        print(f"[!] Aucun fichier de production trouvé.")
        print(f"    Vérité   : {data_root}/FORM*/EXAM_FORM*_NNNNN.xlsx")
        print(f"    Prod     : {results_root}/EXAM_FORM*_RESULTS/EXAM_FORM*_NNNNN.xlsx")
        print(f"    -> Lance d'abord 'python main.py' pour générer les xlsx de prod.")
        return

    # Stats par axe
    axis_stats = {}  # axis -> [ok, total]
    csv_rows = [["form", "id", "sheet", "field_or_cell", "axis",
                 "truth", "prod", "ok"]]

    for form, sid, tp, pp in pairs:
        twb = openpyxl.load_workbook(tp, data_only=True)
        pwb = openpyxl.load_workbook(pp, data_only=True)
        # PAGE-01
        if "PAGE-01" in twb.sheetnames and "PAGE-01" in pwb.sheetnames:
            for r, label, axis, t, p, ok in compare_page1(pwb["PAGE-01"],
                                                          twb["PAGE-01"]):
                axis_stats.setdefault(axis, [0, 0])
                axis_stats[axis][1] += 1
                if ok: axis_stats[axis][0] += 1
                csv_rows.append([form, sid, "PAGE-01", label, axis,
                                 str(t), str(p), int(ok)])
        # EXAM
        if "EXAM" in twb.sheetnames and "EXAM" in pwb.sheetnames:
            for cell, axis, t, p, ok in compare_exam(pwb["EXAM"], twb["EXAM"]):
                axis_stats.setdefault(axis, [0, 0])
                axis_stats[axis][1] += 1
                if ok: axis_stats[axis][0] += 1
                csv_rows.append([form, sid, "EXAM", cell, axis,
                                 str(t), str(p), int(ok)])

    out_csv = "compare_results.csv"
    with open(out_csv, "w", newline="") as f:
        csv.writer(f).writerows(csv_rows)

    print(f"\n=== Comparaison vérité vs production ({len(pairs)} xlsx) ===\n")
    print(f"{'AXE':<12}{'OK':>6}{'TOTAL':>8}{'ACC':>10}")
    print("-" * 36)
    for axis in ("imprime", "manuscrit", "graphique", "signature", "autre"):
        if axis not in axis_stats: continue
        ok, tot = axis_stats[axis]
        print(f"{axis:<12}{ok:>6}{tot:>8}{100*ok/max(1,tot):>9.1f}%")
    print("-" * 36)
    total_ok = sum(v[0] for v in axis_stats.values())
    total    = sum(v[1] for v in axis_stats.values())
    print(f"{'GLOBAL':<12}{total_ok:>6}{total:>8}{100*total_ok/max(1,total):>9.1f}%")
    print(f"\nDétails -> {out_csv}")


if __name__ == "__main__":
    main()
